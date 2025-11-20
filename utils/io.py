import torch
import os
import imageio
import numpy as np
from skimage import img_as_ubyte
from rich.progress import track
import cv2
from openpyxl import Workbook, load_workbook

def is_video_file(path):
    return path.lower().endswith(('.mp4', '.mov', '.avi', '.mkv'))

def compute_bitrate(bits, fps,frames):
    return ((bits*fps)/(1000*frames))

def detach_frame(x):
    x = torch.squeeze(x, dim=0).data.cpu().numpy()
    return img_as_ubyte(np.transpose(x, [1, 2, 0]))

def load_video(video_info, n_frames=-1):
    cap = cv2.VideoCapture(video_info)
    
    if not cap.isOpened():
        print(f"Error: Could not open video {video_info}")
        return []

    frames = []
    idx = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        if 0 < n_frames <= idx:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        frames.append(rgb_frame)
        idx += 1
    
    cap.release()
    return frames

def images2video(images, wfp, **kwargs):
    fps = kwargs.get('fps', 30)
    video_format = kwargs.get('format', 'mp4')  # default is mp4 format
    codec = kwargs.get('codec', 'libx264')  # default is libx264 encoding
    quality = kwargs.get('quality')  # video quality
    pixelformat = kwargs.get('pixelformat', 'yuv420p')  # video pixel format
    image_mode = kwargs.get('image_mode', 'rgb')
    macro_block_size = kwargs.get('macro_block_size', 2)
    ffmpeg_params = ['-crf', str(kwargs.get('crf', 23))]

    writer = imageio.get_writer(
        wfp, fps=fps, format=video_format,
        codec=codec, quality=quality, ffmpeg_params=ffmpeg_params, pixelformat=pixelformat, macro_block_size=macro_block_size
    )

    n = len(images)
    for i in track(range(n), description='Writing', transient=True):
        if image_mode.lower() == 'bgr':
            writer.append_data(images[i][..., ::-1])
        else:
            writer.append_data(images[i])

    writer.close()

def write_results(results, filename="results.xlsx", framewise=False):
    if framewise:
        max_len = max(len(v) if isinstance(v, list) else 1 for v in results.values())
        expanded_results = {}
        for key, val in results.items():
            if isinstance(val, list):
                expanded_results[key] = val
            else:
                expanded_results[key] = [val] * max_len
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(expanded_results.keys()))
        for i in range(max_len):
            row = [expanded_results[key][i] for key in expanded_results]
            ws.append(row)
    else:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            next_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(results.keys()))
            next_row = 2
        values = list(results.values())
        for col_index, value in enumerate(values, start=1):
            ws.cell(row=next_row, column=col_index, value=value)
    wb.save(filename)